import openpyxl
from collections import Counter

wb = openpyxl.load_workbook(
    r'c:\Users\kc10g\OneDrive\Documents\JCD Enterprises\Final Validation Inventory File 23FEB2026 (V2X)-OA LIST.xlsx',
    data_only=True
)
ws = wb['Master Inventory Sheet']

EXCLUDE = {'A', 'B', 'C', 'D'}

# Collect ALL unique location values across all non-ABCD rows, grouped by site
site_locations = {}

for row in ws.iter_rows(min_row=4, max_row=ws.max_row, values_only=True):
    site = row[1]
    part = row[3]
    location = row[5]
    qty = row[7]
    cond = row[9]

    if part is None or cond is None:
        continue
    cond_str = str(cond).strip().upper()
    if cond_str in EXCLUDE:
        continue

    site_str = str(site).strip() if site else 'UNKNOWN'
    loc_str = str(location).strip() if location else 'BLANK'

    if site_str not in site_locations:
        site_locations[site_str] = Counter()
    site_locations[site_str][loc_str] += 1

wb.close()

# Look for locations that might indicate "out for repair"
repair_keywords = ['repair', 'vendor', 'jls', 'otr', 'out', 'rpr', 'sent', 'ship']

print('ALL LOCATIONS THAT MAY INDICATE "OUT FOR REPAIR":')
print('(searching for keywords: repair, vendor, jls, otr, out, rpr, sent, ship)')
print('=' * 80)
for site in sorted(site_locations.keys()):
    found = []
    for loc, count in site_locations[site].most_common():
        loc_lower = loc.lower()
        if any(kw in loc_lower for kw in repair_keywords):
            found.append((loc, count))
    if found:
        print(f'\nSite {site}:')
        for loc, count in found:
            print(f'  "{loc}": {count} rows')

print('\n\n')
print('ALL UNIQUE LOCATIONS BY SITE (full list):')
print('=' * 80)
for site in sorted(site_locations.keys()):
    locs = site_locations[site]
    print(f'\nSite {site} ({sum(locs.values())} total rows, {len(locs)} unique locations):')
    for loc, count in locs.most_common():
        print(f'  "{loc}" : {count}')
